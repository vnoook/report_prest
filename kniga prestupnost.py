# Программа для создания отчёта Преступность.
# В папке с программой должен быть пустой шаблон Преступность и файл из ИЦ.
# Программа пропускает 3 строки из переменной 'rows_for_skip'.
# После отработки программы остаётся переименовать файл под период,
#    поменять жёлтые места на 1, 2 и последнем листе период.

import os
import time
import datetime
import openpyxl

# считаю время скрипта
time_start = time.time()
print('начинается ' + '.'*19)


# функция для анализа что выдавать в ячейку
# в ячейке может быть целое число, дробное число, строка или пусто
def conv_cell(cell_value):
    if type(cell_value) == int:
        return cell_value
    elif type(cell_value) == float:
        return cell_value
    else:
        if cell_value == '***' or cell_value == '0,0':
            return cell_value
        elif cell_value is None:
            cell_value = ''
            return cell_value
        else:
            cell_value = float(cell_value.replace(',', '.'))
            return cell_value


# функция для анализа что в ячейке
# если в ячейке то, что можно преобразовать в число, то выдать иначе выдать False
def int_cell(cell_value):
    if type(cell_value) == int:
        return cell_value
    else:
        return False


# функция составления названия файла для сохранения
# состоит из названия отчёта + месяца + года
def name_of_file():
    # текущие месяц и год
    number_of_month = datetime.datetime.today().month
    number_of_year = datetime.datetime.today().year

    # если запустили в январе, то (месяц и год) надо сменить на (декабрь и (год-1))
    # иначе (месяц-1)
    if number_of_month == 1:
        number_of_month = 12
        number_of_year -= 1
    else:
        number_of_month -= 1

    # если номер месяца цифра, то добавить 0 в начало
    # иначе просто перевести в строку
    if number_of_month < 10:
        name_month = '0'+str(number_of_month)
    else:
        name_month = str(number_of_month)

    file_name = 'Преступность-' + name_month + '-' + str(number_of_year) + '.xlsx'
    return file_name


# функция составления названия файла для сохранения
# состоит из года + месяц
def name_of_file_ic():
    # текущие месяц и год
    number_of_month = datetime.datetime.today().month
    number_of_year = datetime.datetime.today().year

    name_year = str(number_of_year)
    if number_of_month < 10:
        name_month = '0'+str(number_of_month)
    else:
        name_month = str(number_of_month)

    str_period = name_year + '-' + name_month + '.xlsx'
    return str_period


# соответствие листов и ячеек с которых начинаются данные в файле из ИЦ
xl_cell_begin = {
                '2': 'B5:F32',
                '3': 'B8:O57',
                '4': 'B8:L57',
                '5': 'B9:O58',
                '6': 'B8:O57',
                '7': 'B8:O57',
                '8': 'B8:O57',
                '9': 'B8:O57',
                '10': 'B8:O57',
                '11': 'B8:O57',
                '12': 'B8:O57',
                '13': 'D6:O55',
                '14': 'D6:O55',
                '15': 'B10:Y59',
                '16': 'B6:U55',
                '17': 'B9:J58',
                '18': 'B7:N56',
                '19': 'B7:N56',
                '20': 'B7:M56',
                '21': 'B7:N56',
                '22': 'B7:N56',
                '23': 'B9:M58',
                '24': 'B9:M58',
                '25': 'B7:AB56',
                '26': 'B8:U57',
                '27': 'B8:S57',
                '28': 'B8:G57'
                }

# файлы для работы
xl_template = 'ШАБЛОН Преступность-01-2021.xlsx'
xl_source = name_of_file_ic()

if not os.path.exists(xl_source):
    print(f'\n')
    print(f'Ожидается файл {xl_source}, его нет в рабочей папке!\n')
    print(f'Скопируйте файл и перезапустите программу!\n')
    input(f'Нажмите ENTER')
    exit()

# строки для пропуска, только в 1ЕМ
rows_for_skip = (11, 12, 49)

# открываю книгу шаблон в которую вставляю данные
# wb_prestup - файл шаблона
wb_prestup = openpyxl.load_workbook(xl_template)

# открываю книгу из которой беру данные
# wb_file_data - файл шаблона
wb_file_data = openpyxl.load_workbook(xl_source)

# иду по листам шаблона чтобы вставить данные из файла
# беру все листы шаблона по очереди
for i_list in xl_cell_begin:
    # назначаю в файлах активный лист из списка xl_cell_begin
    # wb_prestup_s - лист в шаблоне, wb_file_data_s - лист в шаблоне
    wb_prestup_s = wb_prestup[str(i_list)]
    wb_file_data_s = wb_file_data[str(i_list)]

    # прохожу по всем листам
    # если лист 2, то спец обработка, если другие, то считаю все одинаково
    if i_list == '2':
        # надо обработать 2 диапазона
        # B5:F25  -> R5C2:R25C6
        # B27:F32 -> R27C2:R32C6

        # беру строки
        for i_rows in range(5, 32+1):
            # пропускаю 26 строку
            if i_rows == 26:
                pass
            # беру колонки
            else:
                for i_col in range(2, 6+1):
                    # тут берётся ячейка и вставляется в ячейку со сдвигом колонки '+1' или без
                    if i_col == 2 or i_col == 3:
                        wb_prestup_s.cell(i_rows, i_col).value = conv_cell(wb_file_data_s.cell(i_rows, i_col).value)
                    else:
                        wb_prestup_s.cell(i_rows, i_col+1).value = conv_cell(wb_file_data_s.cell(i_rows, i_col).value)
    else:
        # выбираю диапазон ячеек из источника данных - тип данных кортеж кортежей
        # можно было бы обратиться к данным и так cells_range[0][0].value
        cells_range = wb_file_data_s[
                                    xl_cell_begin[str(i_list)].split(':')[0]:
                                    xl_cell_begin[str(i_list)].split(':')[1]
                                    ]

        # переменная сдвига при пропуске из rows_for_skip
        i_shift = 0

        # беру построчно из кортежа диапазона
        for range_rows in cells_range:
            # проверяю какая строка и если надо пропустить, то пропускаю и увеличиваю сдвиг i_shift
            # тут cells_range.index(range_rows)+1 потому, что кортеж начинается с 0
            if cells_range.index(range_rows)+1 in rows_for_skip:
                i_shift += 1
            else:
                # тут range_rows.index(range_cols)+1 потому, что кортеж начинается с 0
                for range_cols in range_rows:
                    # тут берётся ячейка с диапазона и вставляется в ячейку со сдвигами '+7' и '+1'
                    wb_prestup_s.cell(
                                     (cells_range.index(range_rows)+1)+7-i_shift,
                                     (range_rows.index(range_cols)+1)+1
                                     ).value = conv_cell(range_cols.value)

# закрываю файл из которого беру данные
wb_file_data.close()

# сохраняю файл шаблона и закрываю его
wb_prestup.save(name_of_file())
wb_prestup.close()

# считаю время скрипта
time_finish = time.time()
print('\n' + '.'*30 + ' закончено за', round(time_finish-time_start, 3), 'секунд')

# закрываю программу
input('Нажмите ENTER')
